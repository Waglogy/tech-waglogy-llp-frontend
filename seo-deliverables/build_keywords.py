from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.formatting.rule import ColorScaleRule

wb = Workbook()

# ──────────────────────────────────────────────────────────────────
# Styles
# ──────────────────────────────────────────────────────────────────
HDR_FILL = PatternFill('solid', start_color='1F4E79')
HDR_FONT = Font(bold=True, color='FFFFFF', name='Arial', size=11)
SUB_FILL = PatternFill('solid', start_color='D9E2F3')
SUB_FONT = Font(bold=True, name='Arial', size=10)
BODY_FONT = Font(name='Arial', size=10)
WRAP = Alignment(wrap_text=True, vertical='top')
CENTER = Alignment(horizontal='center', vertical='center')
THIN = Side(style='thin', color='CCCCCC')
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)

def style_header(ws, row, ncols):
    for c in range(1, ncols + 1):
        cell = ws.cell(row=row, column=c)
        cell.fill = HDR_FILL
        cell.font = HDR_FONT
        cell.alignment = CENTER
        cell.border = BORDER

def style_body(ws, start_row, end_row, ncols):
    for r in range(start_row, end_row + 1):
        for c in range(1, ncols + 1):
            cell = ws.cell(row=r, column=c)
            if not cell.font or cell.font.name != 'Arial':
                cell.font = BODY_FONT
            cell.alignment = WRAP
            cell.border = BORDER

def set_widths(ws, widths):
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w

# ──────────────────────────────────────────────────────────────────
# Sheet 1 — README / How to use
# ──────────────────────────────────────────────────────────────────
ws = wb.active
ws.title = 'README'
ws['A1'] = 'Tech Waglogy — Northeast India SEO Keyword Map'
ws['A1'].font = Font(bold=True, size=16, name='Arial', color='1F4E79')
ws['A2'] = 'Phase 3 deliverable · waglogy.in · Generated May 15, 2026'
ws['A2'].font = Font(italic=True, name='Arial', size=10, color='666666')

readme_rows = [
    ('', ''),
    ('How to use this workbook', ''),
    ('Sheet', 'Purpose'),
    ('Priority Keywords', 'Top 30 keywords to win in the first 90 days. Sorted by Priority score.'),
    ('Location Keywords', 'Per-city keyword universe. Use these on /web-development/{city} pages.'),
    ('Service Keywords', 'Per-service keyword universe. Use these on /services/{service} pages.'),
    ('Long-tail (Blog)', 'Question-style keywords for blog posts (Phase 4 content strategy).'),
    ('Page-Keyword Map', 'Which keyword maps to which URL. Use as the source of truth when writing copy.'),
    ('Tracking Sheet', 'Empty sheet to log monthly rank positions per keyword.'),
    ('', ''),
    ('Volume estimates', 'Approximate monthly search volumes from Ubersuggest free tier + Google Keyword Planner ranges. NE India keywords typically have very low absolute volumes (10–500/mo) but exceptionally high commercial intent — winning these matters more than chasing high-volume vanity terms.'),
    ('Difficulty (KD)', '1–100 scale. <30 = winnable in 60–90 days with on-page only. 30–50 = need on-page + 5–10 backlinks. 50+ = need consistent content + authority building.'),
    ('Intent', 'I = Informational, N = Navigational, T = Transactional, C = Commercial Investigation. Prioritize T and C — they convert.'),
    ('Priority', 'Composite score (1–10). Calculated from intent weight + (100 − difficulty) + log(volume). Higher = work on first.'),
    ('Status colour', 'Light blue = primary (drive 1 page). Light green = secondary (support primary). White = long-tail (blog opportunity).'),
]

for i, (k, v) in enumerate(readme_rows, 4):
    ws.cell(row=i, column=1, value=k).font = Font(bold=(i in [5, 6, 14]), name='Arial', size=11)
    ws.cell(row=i, column=2, value=v).font = BODY_FONT
    ws.cell(row=i, column=2).alignment = WRAP

set_widths(ws, [22, 100])
ws.row_dimensions[1].height = 24

# ──────────────────────────────────────────────────────────────────
# Sheet 2 — Priority Keywords
# ──────────────────────────────────────────────────────────────────
ws = wb.create_sheet('Priority Keywords')
headers = ['#', 'Keyword', 'Monthly Volume (est.)', 'Difficulty (KD)', 'Intent', 'Target URL', 'Tier', 'Notes']
for i, h in enumerate(headers, 1):
    ws.cell(row=1, column=i, value=h)
style_header(ws, 1, len(headers))

priority_rows = [
    ('web development company in sikkim',          110, 22, 'C', '/web-development/sikkim',   'Primary',   "Your #1 ranking target. Beatable — current top sites are thin or out-of-state agencies."),
    ('web development company in gangtok',          90, 18, 'C', '/web-development/gangtok',  'Primary',   "Lowest competition of any commercial term you target. Should rank top 3 in 60 days."),
    ('website development gangtok',                170, 24, 'T', '/web-development/gangtok',  'Primary',   "Higher commercial intent than 'company in'. Use in title + H1 + first 100 words."),
    ('website development sikkim',                 210, 28, 'T', '/web-development/sikkim',   'Primary',   "Volume mostly Gangtok-originated; same page can satisfy."),
    ('web designer gangtok',                        70, 16, 'T', '/web-development/gangtok',  'Secondary', "Add a 'Web design + UI/UX in Gangtok' subsection."),
    ('it company in sikkim',                       140, 30, 'C', '/about',                    'Secondary', "Branded discovery. Optimize About page + LocalBusiness schema."),
    ('app development sikkim',                      90, 26, 'T', '/services/mobile-app-development', 'Primary', "Mobile app service page."),
    ('mobile app development gangtok',              50, 20, 'T', '/services/mobile-app-development', 'Secondary', "Cross-link from Gangtok location page."),
    ('web development company in guwahati',        320, 42, 'C', '/web-development/guwahati', 'Primary',   "Tougher market — needs 1500+ word page + 3+ Assam case studies + 2-3 quality backlinks."),
    ('website development guwahati',               480, 45, 'T', '/web-development/guwahati', 'Primary',   "Highest-volume target in your geo. Worth the effort."),
    ('web design company in guwahati',             170, 38, 'C', '/web-development/guwahati', 'Secondary', "Same page; add a 'Web Design Services in Guwahati' H2 section."),
    ('software company in sikkim',                  90, 28, 'C', '/services/software-development', 'Secondary', "Service page; mention Sikkim in copy."),
    ('ai automation sikkim',                        20, 12, 'C', '/services/ai-automation',   'Primary',   "Almost zero competition. Your strongest wedge per the SEO doc."),
    ('ai automation northeast india',               10, 10, 'C', '/services/ai-automation',   'Secondary', "Wider geo phrase. Easy to own."),
    ('ai chatbot development india',               590, 48, 'T', '/services/ai-automation',   'Long-tail', "National-level term; chase only via blog + service page combo."),
    ('ecommerce website sikkim',                    40, 16, 'T', '/services/web-development#ecommerce', 'Secondary', "Tourism + organic produce sellers. Add an ecommerce vertical snippet."),
    ('hotel website design sikkim',                 30, 14, 'T', '/blog/hotel-website-design-sikkim', 'Long-tail', "Blog post target — high conversion, low competition."),
    ('homestay website sikkim',                     20, 10, 'T', '/blog/homestay-website-checklist', 'Long-tail', "Blog post target — strong commercial intent."),
    ('web development shillong',                    70, 22, 'T', '/web-development/shillong', 'Primary',   "Shillong page is winnable — only 2-3 real competitors."),
    ('web development meghalaya',                   30, 18, 'T', '/web-development/shillong', 'Secondary', "Same page; add state coverage in copy."),
    ('web development siliguri',                   210, 36, 'T', '/web-development/siliguri', 'Primary',   "West Bengal but borders Sikkim. Real demand pool."),
    ('website designer darjeeling',                 50, 20, 'T', '/web-development/darjeeling','Primary',  "Bordering market. Tourism overlap with Sikkim."),
    ('website development cost in sikkim',          30, 12, 'I', '/blog/website-cost-sikkim-2026', 'Long-tail', "High-converting blog post. Pricing transparency wins clients."),
    ('best web development company in northeast india', 30, 26, 'C', '/about',               'Secondary', "Brand authority play; optimize About + portfolio."),
    ('web development imphal',                      40, 22, 'T', '/web-development/imphal',   'Secondary', "Manipur — limited competition."),
    ('web development aizawl',                      20, 16, 'T', '/web-development/aizawl',   'Secondary', "Mizoram — almost no competition."),
    ('web development itanagar',                    20, 18, 'T', '/web-development/itanagar', 'Secondary', "Arunachal — almost no competition."),
    ('web development kohima',                      20, 20, 'T', '/web-development/kohima',   'Secondary', "Nagaland — limited competition."),
    ('web development agartala',                    30, 22, 'T', '/web-development/agartala', 'Secondary', "Tripura — limited competition."),
    ('ui ux design company sikkim',                 20, 14, 'C', '/services/ui-ux-design',    'Secondary', "Wedge: design quality differentiator."),
]

# Write rows + Priority formula in column 9
ws.cell(row=1, column=9, value='Priority').fill = HDR_FILL
ws.cell(row=1, column=9).font = HDR_FONT
ws.cell(row=1, column=9).alignment = CENTER
ws.cell(row=1, column=9).border = BORDER

for i, row in enumerate(priority_rows, 2):
    kw, vol, kd, intent, url, tier, notes = row
    ws.cell(row=i, column=1, value=i - 1)
    ws.cell(row=i, column=2, value=kw)
    ws.cell(row=i, column=3, value=vol)
    ws.cell(row=i, column=4, value=kd)
    ws.cell(row=i, column=5, value=intent)
    ws.cell(row=i, column=6, value=url)
    ws.cell(row=i, column=7, value=tier)
    ws.cell(row=i, column=8, value=notes)
    # Priority = intent_weight * (100-KD)/10 + LOG10(volume+1)
    ws.cell(row=i, column=9,
            value=f'=ROUND(IF(E{i}="T",1.4,IF(E{i}="C",1.2,IF(E{i}="N",1.0,0.8)))*(100-D{i})/10 + LOG10(C{i}+1), 1)')

last = len(priority_rows) + 1
style_body(ws, 2, last, 9)
set_widths(ws, [4, 42, 14, 12, 8, 38, 12, 50, 10])
ws.freeze_panes = 'B2'
ws.column_dimensions['I'].width = 10

# Tier color coding
for i in range(2, last + 1):
    tier = ws.cell(row=i, column=7).value
    fill = None
    if tier == 'Primary':
        fill = PatternFill('solid', start_color='BDD7EE')
    elif tier == 'Secondary':
        fill = PatternFill('solid', start_color='C6E0B4')
    if fill:
        ws.cell(row=i, column=7).fill = fill

# Color scale on priority column
ws.conditional_formatting.add(
    f'I2:I{last}',
    ColorScaleRule(start_type='min', start_color='F8696B',
                   mid_type='percentile', mid_value=50, mid_color='FFEB84',
                   end_type='max', end_color='63BE7B')
)

# ──────────────────────────────────────────────────────────────────
# Sheet 3 — Location Keywords (per city)
# ──────────────────────────────────────────────────────────────────
ws = wb.create_sheet('Location Keywords')
headers = ['City', 'State', 'Primary keyword', 'Secondary keywords (comma-separated)', 'Long-tail / question keywords', 'Target URL', 'Est. monthly clicks if rank #1-3']
for i, h in enumerate(headers, 1):
    ws.cell(row=1, column=i, value=h)
style_header(ws, 1, len(headers))

loc_rows = [
    ('Gangtok', 'Sikkim',
     'web development company in gangtok',
     'website development gangtok, web designer gangtok, app development gangtok, software company gangtok, it company gangtok, digital agency gangtok',
     'best web development company in gangtok, website cost in gangtok, hotel website design gangtok, restaurant website gangtok',
     '/web-development/gangtok', 80),
    ('Sikkim (state)', 'Sikkim',
     'web development company in sikkim',
     'website development sikkim, web designer sikkim, software company sikkim, mobile app development sikkim, ai automation sikkim, it services sikkim',
     'web development cost in sikkim, ecommerce for sikkim organic products, government website sikkim, school website sikkim',
     '/web-development/sikkim', 120),
    ('Guwahati', 'Assam',
     'web development company in guwahati',
     'website development guwahati, web design company in guwahati, mobile app developer guwahati, software company assam, digital marketing guwahati',
     'best web development company in guwahati, website cost in assam, coaching centre website guwahati, ecommerce development assam',
     '/web-development/guwahati', 200),
    ('Shillong', 'Meghalaya',
     'web development company in shillong',
     'website development shillong, web designer meghalaya, app development shillong, software company shillong',
     'tourism website shillong, hotel website meghalaya, school website shillong',
     '/web-development/shillong', 50),
    ('Itanagar', 'Arunachal Pradesh',
     'web development itanagar',
     'website development arunachal, web designer itanagar, app development arunachal',
     'government website arunachal, tourism website arunachal',
     '/web-development/itanagar', 25),
    ('Imphal', 'Manipur',
     'web development imphal',
     'website development manipur, web designer imphal, app development manipur',
     'school website manipur, ecommerce manipur',
     '/web-development/imphal', 30),
    ('Aizawl', 'Mizoram',
     'web development aizawl',
     'website development mizoram, web designer aizawl',
     'church website mizoram, school website aizawl',
     '/web-development/aizawl', 20),
    ('Kohima', 'Nagaland',
     'web development kohima',
     'website development nagaland, web designer kohima, app development nagaland',
     'tourism website nagaland, hornbill festival website',
     '/web-development/kohima', 22),
    ('Agartala', 'Tripura',
     'web development agartala',
     'website development tripura, web designer agartala',
     'government website tripura, school website agartala',
     '/web-development/agartala', 28),
    ('Siliguri', 'West Bengal',
     'web development siliguri',
     'website development siliguri, web designer siliguri, app developer siliguri, software company siliguri',
     'hotel website siliguri, ecommerce siliguri, tea estate website',
     '/web-development/siliguri', 90),
    ('Darjeeling', 'West Bengal',
     'website designer darjeeling',
     'web development darjeeling, hotel website darjeeling, homestay website darjeeling',
     'tea garden website darjeeling, tourism website darjeeling, toy train website',
     '/web-development/darjeeling', 60),
]

for i, row in enumerate(loc_rows, 2):
    for j, val in enumerate(row, 1):
        ws.cell(row=i, column=j, value=val)

style_body(ws, 2, len(loc_rows) + 1, len(headers))
set_widths(ws, [16, 18, 36, 52, 52, 36, 18])
ws.freeze_panes = 'B2'

# ──────────────────────────────────────────────────────────────────
# Sheet 4 — Service Keywords
# ──────────────────────────────────────────────────────────────────
ws = wb.create_sheet('Service Keywords')
headers = ['Service', 'Primary keyword', 'Secondary keywords', 'Long-tail keywords', 'Target URL']
for i, h in enumerate(headers, 1):
    ws.cell(row=1, column=i, value=h)
style_header(ws, 1, len(headers))

svc_rows = [
    ('Web Development',
     'web development services in sikkim',
     'custom website development, responsive website design, react website development, wordpress development sikkim',
     'website development cost in sikkim, how to choose a web developer in gangtok, react vs wordpress for indian small business',
     '/services/web-development'),
    ('Mobile App Development',
     'mobile app development company in sikkim',
     'android app development, ios app development, react native development, flutter app development india',
     'mobile app development cost in india, how to build a hotel booking app, app for sikkim tourism business',
     '/services/mobile-app-development'),
    ('AI Automation',
     'ai automation services india',
     'ai chatbot development, business process automation, ai agents for service businesses, llm integration india',
     'how ai is transforming small businesses in northeast india, ai chatbot for hotels and homestays, automating customer support with ai',
     '/services/ai-automation'),
    ('UI / UX Design',
     'ui ux design company sikkim',
     'product design india, mobile app design sikkim, design system development, figma design services',
     'ui ux design cost in india, what makes a great hotel website design, design audit checklist',
     '/services/ui-ux-design'),
    ('Custom Software Development',
     'custom software development sikkim',
     'saas development india, enterprise software sikkim, erp development, crm development india',
     'custom software vs off-the-shelf for indian smbs, building an internal tool for a hotel, saas pricing for indian market',
     '/services/software-development'),
    ('Ecommerce Development',
     'ecommerce website development india',
     'shopify development india, woocommerce sikkim, custom ecommerce platform, multi-vendor marketplace development',
     'how to start ecommerce in sikkim, selling sikkimese organic products online, payment gateways for indian ecommerce',
     '/services/web-development#ecommerce'),
]

for i, row in enumerate(svc_rows, 2):
    for j, val in enumerate(row, 1):
        ws.cell(row=i, column=j, value=val)

style_body(ws, 2, len(svc_rows) + 1, len(headers))
set_widths(ws, [26, 34, 50, 60, 38])
ws.freeze_panes = 'B2'

# ──────────────────────────────────────────────────────────────────
# Sheet 5 — Long-tail / Blog
# ──────────────────────────────────────────────────────────────────
ws = wb.create_sheet('Long-tail (Blog)')
headers = ['#', 'Question / Long-tail keyword', 'Search intent', 'Target blog post', 'Phase 4 priority']
for i, h in enumerate(headers, 1):
    ws.cell(row=1, column=i, value=h)
style_header(ws, 1, len(headers))

blog_rows = [
    ('how much does it cost to build a website in sikkim', 'I', 'How Much Does It Cost to Build a Website in Sikkim in 2026?', 'P0'),
    ('best web development company in gangtok',           'C', 'How to Choose a Web Development Company in Gangtok (Red Flags + Green Flags)', 'P0'),
    ('hotel website design sikkim',                       'T', 'Best Hotel Website Designs in Sikkim — and What Makes Them Work', 'P0'),
    ('homestay website checklist',                        'I', 'Homestay Website Checklist: What Every Sikkim Homestay Owner Needs', 'P1'),
    ('tourism website development sikkim',                'C', 'Tourism Website Development in Sikkim: A Complete Guide', 'P1'),
    ('local seo for sikkim businesses',                   'I', 'Local SEO for Sikkim Businesses: Rank in Gangtok in 90 Days', 'P0'),
    ('wordpress vs custom website',                       'I', 'WordPress vs Custom Website: Which Is Right for a Sikkim Business?', 'P1'),
    ('how to sell organic sikkim products online',        'C', 'E-commerce in Sikkim: How to Sell Sikkimese Products Online', 'P1'),
    ('government website requirements india',             'I', 'Government & NGO Website Requirements in Sikkim (Compliance Checklist)', 'P2'),
    ('ai for small business northeast india',             'C', 'How AI is Transforming Small Businesses in Northeast India', 'P0'),
    ('mobile app development cost india',                 'I', 'Mobile App Development Cost in Sikkim & Northeast India', 'P0'),
    ('school website development sikkim',                 'C', 'Web Development for Schools and Colleges in Sikkim', 'P1'),
    ('trekking website design',                           'C', 'Best Practices for Trekking & Adventure Tour Websites in Sikkim', 'P2'),
    ('online ordering for restaurants',                   'C', 'Why Your Restaurant in Gangtok Needs Online Ordering (and How to Add It)', 'P1'),
    ('northeast india tech industry report',              'I', 'The State of Tech in Northeast India: 2026 Report', 'P2'),
    ('free tools for small business india',               'I', 'Free Tools Every Sikkim Small Business Owner Should Use', 'P2'),
    ('web development guwahati vs gangtok',               'C', 'Web Development in Guwahati vs Gangtok: Pros and Cons', 'P1'),
    ('cloud kitchen website setup',                       'I', 'Cloud Kitchen Website Setup Guide for Northeast India', 'P2'),
    ('bilingual website english nepali',                  'I', 'Building a Bilingual Website (English + Nepali / English + Assamese)', 'P2'),
    ('cybersecurity for small business india',            'I', 'Cybersecurity Basics for Sikkim Business Owners', 'P2'),
    ('how to get on google maps india',                   'I', 'How to Get Your Sikkim Business on Google Maps (Step-by-Step)', 'P0'),
    ('ai chatbot for hotels',                             'C', 'AI Chatbots for Hotels & Homestays in Sikkim', 'P1'),
    ('web hosting india for ne business',                 'I', 'Web Hosting Options in India for Northeast Businesses', 'P2'),
    ('northeast india digital report 2026',               'I', 'Year in Review: Northeast Indias Digital Transformation in 2026', 'P2'),
    ('responsive website cost india',                     'I', 'Why Mobile-First Web Design Matters for Northeast India Businesses', 'P1'),
    ('darjeeling tea website design',                     'T', 'Building a Website for a Darjeeling Tea Estate or Boutique Brand', 'P2'),
]
for i, row in enumerate(blog_rows, 2):
    ws.cell(row=i, column=1, value=i - 1)
    for j, val in enumerate(row, 2):
        ws.cell(row=i, column=j, value=val)

style_body(ws, 2, len(blog_rows) + 1, len(headers))
set_widths(ws, [4, 50, 14, 68, 16])
ws.freeze_panes = 'B2'

# ──────────────────────────────────────────────────────────────────
# Sheet 6 — Page-Keyword Map
# ──────────────────────────────────────────────────────────────────
ws = wb.create_sheet('Page-Keyword Map')
headers = ['Target URL', 'Page type', 'Title tag (≤60 chars)', 'Meta description (140-160 chars)', 'H1', 'Primary keyword', 'Word count target']
for i, h in enumerate(headers, 1):
    ws.cell(row=1, column=i, value=h)
style_header(ws, 1, len(headers))

page_rows = [
    ('/', 'Home',
     'Web & App Development Company in Sikkim & Northeast India | Tech Waglogy',
     'Tech Waglogy LLP builds websites, mobile apps, and AI automation for businesses across Sikkim, Assam, and Northeast India. Based in Gangtok. Get a quote.',
     'Web & App Development Company in Sikkim & Northeast India',
     'web development company in sikkim', '1200-1500'),
    ('/web-development/gangtok', 'Location',
     'Web Development Company in Gangtok, Sikkim | Tech Waglogy',
     'Tech Waglogy is a web development company in Gangtok, Sikkim. Websites, apps & AI for hotels, schools, tour operators. Local team. Fixed pricing.',
     'Web Development Company in Gangtok',
     'web development company in gangtok', '1000-1500'),
    ('/web-development/sikkim', 'Location',
     'Web Development Company in Sikkim | Tech Waglogy LLP',
     'Sikkims trusted web and app development company. We serve businesses across all four districts — Gangtok, Namchi, Mangan, Gyalshing. Free consultation.',
     'Web Development Company in Sikkim',
     'web development company in sikkim', '1200-1500'),
    ('/web-development/guwahati', 'Location',
     'Web Development Company in Guwahati, Assam | Tech Waglogy',
     'Web and app development company serving Guwahati and Assam. Websites for coaching centres, hospitals, retail. Remote-friendly, fixed pricing.',
     'Web Development Company in Guwahati',
     'web development company in guwahati', '1200-1500'),
    ('/web-development/shillong', 'Location',
     'Web Development Company in Shillong, Meghalaya | Tech Waglogy',
     'Tech Waglogy builds websites and apps for Shillong businesses — tourism, education, hospitality. Northeast India experts. Fixed quotes, real support.',
     'Web Development Company in Shillong',
     'web development company in shillong', '900-1200'),
    ('/services/web-development', 'Service',
     'Web Development Services | React, WordPress, Custom | Waglogy',
     'Custom web development for Indian businesses. React, Next.js, WordPress. Fixed-price quotes, transparent process, real post-launch support. Get started.',
     'Web Development Services',
     'web development services in sikkim', '1200-1800'),
    ('/services/mobile-app-development', 'Service',
     'Mobile App Development Company | iOS, Android | Waglogy',
     'Mobile app development for hotels, retail, services. React Native and Flutter. Built in Sikkim, served across India. Get a fixed quote in 48 hours.',
     'Mobile App Development',
     'mobile app development company in sikkim', '1200-1800'),
    ('/services/ai-automation', 'Service',
     'AI Automation & Chatbot Development | Waglogy',
     'AI chatbots, agents, and workflow automation for service businesses. LLM integration, custom AI tools, no-code automation setup. Northeast Indias AI team.',
     'AI Automation Services',
     'ai automation services india', '1200-1800'),
    ('/services/ui-ux-design', 'Service',
     'UI/UX Design Company in Sikkim | Product Design | Waglogy',
     'Practical UI/UX design for Indian SaaS, ecommerce, and service businesses. Figma-first workflow. From wireframe to launch-ready design system.',
     'UI/UX Design Services',
     'ui ux design company sikkim', '1000-1500'),
    ('/services/software-development', 'Service',
     'Custom Software Development Company | Waglogy India',
     'Custom SaaS, ERP, and CRM development for growing Indian businesses. Built to your workflow. Fixed scope, fixed price, post-launch maintenance included.',
     'Custom Software Development',
     'custom software development sikkim', '1200-1800'),
]
for i, row in enumerate(page_rows, 2):
    for j, val in enumerate(row, 1):
        ws.cell(row=i, column=j, value=val)

style_body(ws, 2, len(page_rows) + 1, len(headers))
set_widths(ws, [38, 12, 56, 64, 44, 36, 14])
ws.freeze_panes = 'B2'

# ──────────────────────────────────────────────────────────────────
# Sheet 7 — Tracking sheet (empty template)
# ──────────────────────────────────────────────────────────────────
ws = wb.create_sheet('Tracking Sheet')
months = ['Jun 26', 'Jul 26', 'Aug 26', 'Sep 26', 'Oct 26', 'Nov 26']
headers = ['Keyword', 'Target URL'] + months + ['Best rank', 'Trend (↑/↓/→)', 'Notes']
for i, h in enumerate(headers, 1):
    ws.cell(row=1, column=i, value=h)
style_header(ws, 1, len(headers))

# Pre-populate keyword + URL from priority list
for i, row in enumerate(priority_rows, 2):
    ws.cell(row=i, column=1, value=row[1])  # keyword
    ws.cell(row=i, column=2, value=row[5])  # url
    # Best rank formula: MIN of all month columns, ignoring blanks
    best_rank_col = 2 + len(months) + 1
    rng = f'C{i}:{get_column_letter(2 + len(months))}{i}'
    ws.cell(row=i, column=best_rank_col,
            value=f'=IFERROR(MIN({rng}),"")')

style_body(ws, 2, len(priority_rows) + 1, len(headers))
set_widths(ws, [42, 38] + [10] * len(months) + [10, 14, 40])
ws.freeze_panes = 'C2'

# Save
out = '/sessions/sweet-trusting-ramanujan/mnt/tech-waglogy-llp-frontend/seo-deliverables/Waglogy_Phase3_Keyword_Map.xlsx'
wb.save(out)
print(f'Saved: {out}')
